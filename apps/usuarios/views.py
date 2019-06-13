from django.http.response import JsonResponse
import openpyxl

#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
from openpyxl.utils import get_column_letter
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.styles import Alignment

#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView
from django.views.generic.edit import CreateView, UpdateView
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from apps.usuarios.utilities import generar_pdf_usuarios
from django.urls import reverse_lazy
from .forms import *
from django.contrib.auth.hashers import make_password
from django.core.mail import send_mail, EmailMessage
from django.core import serializers
import json
import os
#import locale
from apps.usuarios.reports import *
from apps.carrito.views import *

#locale.setlocale(locale.LC_ALL, "esp")
#locale.setlocale(locale.LC_ALL, "es_CO.utf8")

def home(request):
    Usuario.crear_admin()
    schema = connection.schema_name
    usuario = request.user
    tenants = Dominio.objects.exclude(tenant__schema_name='public')
    tenant_data = Tenant.objects.get(schema_name=schema)
    if schema == 'public':
        if request.method == 'POST':
            form = ContactForm(request.POST)
            if form.is_valid():
                msg = EmailMessage('Maravilla Tenants - Solicitud de Información', "Nombre Completo: " +
                                   form.cleaned_data['nombre'] + "<br><br>Correo Electrónico: " +
                                   form.cleaned_data['correo'] + "<br><br>Mensaje: " + form.cleaned_data['mensaje'],
                                   'maravilla.franquicias@gmail.com', ['dianagarco@gmail.com'])
                msg.content_subtype = "html"
                msg.send()
                return redirect('usuarios:home')
        else:
            form = ContactForm()
            return render(request, 'tenants/home_franquicia.html', {'public': tenant_data, 'usuario': usuario, 'form': form})

        return render(request, 'usuarios/home_tenant.html',
                      {'public': tenant_data, 'usuario': usuario, 'tenants': tenants,
                       'form': form})
    else:
        return render(request, 'usuarios/home_tenant.html',
                      {'tenant': tenant_data, 'usuario': usuario, 'orden': get_orden_usuario_pendiente(request)})


def dashboard(request):
    schema = connection.schema_name
    tenant_data = Tenant.objects.get(schema_name=schema)
    usuario = request.user
    mes = datetime.date(datetime.date.today().year, datetime.date.today().month, 1).strftime('%B')
    counter = reportes_counter()
    datos_ventas_diarias = reporte_ventas_diarias()
    datos_productos_vendidos = reporte_productos_vendidos()
    datos_usuarios = reporte_usuarios()
    ventas_anon = reporte_ventas_anuales_anonimos()
    ventas_reg = reporte_ventas_anuales_registrados()
    anual_porcentajes = reporte_ventas_anuales_porcentaje()
    return render(request, 'usuarios/dash_admin.html', {'tenant': tenant_data, 'usuario': usuario, 'mes':mes,
                                                        'datos_ventas_diarias': datos_ventas_diarias, 'counter': counter,
                                                        'datos_productos_vendidos': datos_productos_vendidos,
                                                        'datos_usuarios': datos_usuarios,
                                                        'ventas_anonimos': ventas_anon,
                                                        'ventas_registrados': ventas_reg,
                                                        'ventas_porcentaje': anual_porcentajes}
                                                        )

class DatosActualizar(SuccessMessageMixin, UpdateView):
    model = Usuario
    fields = ['first_name', 'last_name', 'documento', 'ciudad', 'barrio', 'direccion', 'telefono', 'username', 'email']

    success_message = "Datos actualizados exitosamente"
    def get_context_data(self, **kwargs):
        context = super(DatosActualizar, self).get_context_data(**kwargs)
        context['usuario'] = self.request.user
        schema = connection.schema_name
        context['tenant'] = Tenant.objects.get(schema_name=schema)
        if self.request.user.is_authenticated:
            context['orden'] = get_orden_usuario_pendiente(self.request)
        return context

    def get_success_url(self, **kwargs):
        return reverse_lazy("usuarios:home")


class Registro(SuccessMessageMixin, CreateView):
    model = Usuario
    fields = ['first_name', 'last_name', 'documento', 'ciudad', 'barrio', 'direccion', 'telefono', 'username', 'email', 'password']

    success_message = "Gracias por registrarte"
    def form_valid(self, form):
        form_data =form.cleaned_data
        try:
            if form_data['password']:
                form.instance.password= make_password(form_data['password'])
        except KeyError:
            pass
        return super(Registro, self).form_valid(form)

    def get_context_data(self, **kwargs):
        context = super(Registro, self).get_context_data(**kwargs)
        schema = connection.schema_name
        context['tenant'] = Tenant.objects.get(schema_name=schema)
        context['usuario'] = self.request.user
        return context

    def get_success_url(self, **kwargs):
        return reverse_lazy("usuarios:login")


class CrearDigitador(SuccessMessageMixin, CreateView):
    model = Usuario
    fields = ['first_name', 'last_name', 'documento', 'ciudad', 'barrio', 'direccion', 'telefono', 'username', 'email',
              'password']

    success_message = "Digitador creado exitosamente"
    def form_valid(self, form):
        form.instance.is_staff = True
        form_data =form.cleaned_data
        try:
            if form_data['password']:
                form.instance.password= make_password(form_data['password'])
        except KeyError:
            pass

        return super(CrearDigitador, self).form_valid(form)

    def get_context_data(self, **kwargs):
        context = super(CrearDigitador, self).get_context_data(**kwargs)
        context['usuario'] = self.request.user
        return context

    def get_success_url(self, **kwargs):
        return reverse_lazy("usuarios:listar_digitadores")


class DigitadoresListar(ListView):
    model = Usuario

    def get_queryset(self):
        return Usuario.objects.filter(is_superuser=False, is_staff=True)

    def get_context_data(self, **kwargs):
        context = super(DigitadoresListar, self).get_context_data(**kwargs)
        context['usuario'] = self.request.user
        return context


def pdf_usuario(request, staff):
    if staff == 1:
        digitadores = Usuario.objects.filter(is_superuser=False, is_staff=True)
        return generar_pdf_usuarios(digitadores, staff)
    else:
        clientes = Usuario.objects.filter(is_superuser=False, is_staff=False)
        return generar_pdf_usuarios(clientes, staff)


class ClientesListar(ListView):
    model = Usuario

    def get_queryset(self):
        return Usuario.objects.filter(is_superuser=False, is_staff=False)

    def get_context_data(self, **kwargs):
        context = super(ClientesListar, self).get_context_data(**kwargs)
        context['usuario'] = self.request.user
        return context


class UsuarioDetalle(DetailView):
    model = Usuario

    def get_context_data(self, **kwargs):
        context = super(UsuarioDetalle, self).get_context_data(**kwargs)
        context['usuario'] = self.request.user
        return context


def usuario_desactivar(request, id_usuario):
    usuario = Usuario.objects.get(id=id_usuario)
    usuario.is_active = False
    usuario.save()
    if not usuario.is_staff:
        json_data = serializers.serialize('json', [usuario])

        with open("data_id_" + str(usuario.id) + "_" + str(usuario.username) + ".json", "w") as outfile:
            json.dump(json.JSONDecoder().decode(json_data), outfile, indent=4)

        msg = EmailMessage('Maravilla - Eliminación de cuenta', "Tu cuenta ha sido eliminada exitosamente.<br><br>"
                                                                "Muchas gracias por utilizar nuestros servicios y "
                                                                "esperamos que vuelvas de nuevo algún día. :) <br>"
                                                                "Adjunto encontrarás todos tus datos.<br><br>"
                                                                "Hasta Pronto.",
                           'maravilla.franquicias@gmail.com', ['dianagarco@gmail.com'])
        msg.content_subtype = "html"
        msg.attach_file("data_id_" + str(usuario.id) + "_" + str(usuario.username) + ".json")
        msg.send()

        os.remove("data_id_" + str(usuario.id) + "_" + str(usuario.username) + ".json")

    return redirect('usuarios:salir')


def cambiar_contrasena(request):
    if request.method == 'POST':
        form = PasswordChangeForm(data=request.POST, user=request.user)

        if form.is_valid():
            form.save()
            update_session_auth_hash(request, form.user)
            messages.success(request, "La contraseña ha sido actualizada.")
            return redirect(reverse('usuarios:home'))
        else:
            messages.error(request, "Contraseña incorrecta. Intente de nuevo.")
            return redirect(reverse('usuarios:contrasena'))
    else:
        form = PasswordChangeForm(user=request.user)

        return render(request, 'usuarios/contrasena.html', {'usuario': request.user, 'form': form})
#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse

#Nuestra clase hereda de la vista genérica TemplateView
class ReporteClientesExcel(TemplateView):

    #Usamos el método get para generar el archivo excel
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las personas de nuestra base de datos
        personas = Usuario.objects.filter(is_superuser=False, is_staff=False)
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active

        ws.cell(row=2, column=1).alignment  = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=2, column=1).alignment  = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=3, column=1).alignment  = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=3, column=2).alignment  = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=3, column=3).alignment  = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=3, column=4).alignment  = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=3, column=5).alignment  = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=3, column=6).alignment  = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=3, column=7).alignment  = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['A1'] = 'Franquicias Maravilla'
        ws['A2'] = 'REPORTE DE CLIENTES'

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 15



        ws.merge_cells('A1:G1')
        ws.merge_cells('A2:G2')

        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['A3'] = 'ID'
        ws['B3'] = 'USUARIO'
        ws['C3'] = 'DOCUMENTO'
        ws['D3'] = 'NOMBRES'
        ws['E3'] = 'APELLIDOS'
        ws['F3'] = 'CORREO ELECTRONICO'
        ws['G3'] = 'ESTADO'

        cont=4
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for persona in personas:
            ws.cell(row=cont,column=1).value = persona.id
            ws.cell(row=cont,column=2).value = persona.username
            ws.cell(row=cont,column=3).value = persona.documento
            ws.cell(row=cont,column=4).value = persona.first_name
            ws.cell(row=cont,column=5).value = persona.last_name
            ws.cell(row=cont,column=6).value = persona.email
            ws.cell(row=cont,column=7).value = persona.is_active
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteClientesExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReporteDigitadoresExcel(TemplateView):

    #Usamos el método get para generar el archivo excel
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las personas de nuestra base de datos
        personas = Usuario.objects.filter(is_superuser=False, is_staff=True)
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.cell(row=2, column=1).alignment  = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=2, column=1).alignment  = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=3, column=1).alignment  = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=3, column=2).alignment  = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=3, column=3).alignment  = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=3, column=4).alignment  = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=3, column=5).alignment  = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=3, column=6).alignment  = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=3, column=7).alignment  = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)


        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['A1'] = 'Franquicias Maravilla'
        ws['A2'] = 'REPORTE DE DIGITADORES'

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 15



        ws.merge_cells('A1:G1')
        ws.merge_cells('A2:G2')

        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['A3'] = 'ID'
        ws['B3'] = 'USUARIO'
        ws['C3'] = 'DOCUMENTO'
        ws['D3'] = 'NOMBRES'
        ws['E3'] = 'APELLIDOS'
        ws['F3'] = 'CORREO ELECTRONICO'
        ws['G3'] = 'ESTADO'

        cont=4
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for persona in personas:
            ws.cell(row=cont,column=1).value = persona.id
            ws.cell(row=cont,column=2).value = persona.username
            ws.cell(row=cont,column=3).value = persona.documento
            ws.cell(row=cont,column=4).value = persona.first_name
            ws.cell(row=cont,column=5).value = persona.last_name
            ws.cell(row=cont,column=6).value = persona.email
            ws.cell(row=cont,column=7).value = persona.is_active
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteClientesExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

def cambiarEstilo(request):
    if request.is_ajax():
        usuario = request.user
        estilo = request.GET.get('estilo', None)
        try:
            usuario.estilo = str(estilo)
            usuario.save()
            messages.success(request, "El tema ha sido cambiado.")
            return JsonResponse({'success': True})
        except:
            messages.error(request, "Por favor intente de nuevo.")
            return JsonResponse({'estilo': 0})
